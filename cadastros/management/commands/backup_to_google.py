"""
cadastros/management/commands/backup_to_google.py
──────────────────────────────────────────────────
Backup diário do banco (Django dumpdata JSON) + pasta de mídia para o Google Drive.

Variáveis de ambiente obrigatórias (Railway):
    GOOGLE_DRIVE_FOLDER_ID       — ID da pasta de destino no Drive
    GOOGLE_CLIENT_ID             — Client ID OAuth2
    GOOGLE_CLIENT_SECRET         — Client Secret OAuth2
    GOOGLE_REFRESH_TOKEN         — Refresh Token OAuth2 do usuário real

Uso:
    python manage.py backup_to_google
    python manage.py backup_to_google --apenas-db
    python manage.py backup_to_google --apenas-midia
    python manage.py backup_to_google --manter 14    (padrão: 7 dias)
"""
import gzip
import io
import json
import logging
import os
import tarfile

from decouple import config
from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand

logger = logging.getLogger(__name__)

PREFIXO_DB    = 'jr_db_'
PREFIXO_MEDIA = 'jr_media_'
PREFIXO_EXCEL = 'jr_dados_'


class Command(BaseCommand):
    help = 'Backup diário (banco JSON + mídia) para o Google Drive'

    def add_arguments(self, parser):
        parser.add_argument('--apenas-db',    action='store_true')
        parser.add_argument('--apenas-midia', action='store_true')
        parser.add_argument('--apenas-excel', action='store_true')
        parser.add_argument('--manter',       type=int, default=7)

    def handle(self, *args, **options):
        folder           = config('GOOGLE_DRIVE_FOLDER_ID', default='')
        sa_json          = config('GOOGLE_SERVICE_ACCOUNT_JSON', default='')
        client_id        = config('GOOGLE_CLIENT_ID', default='')
        client_secret    = config('GOOGLE_CLIENT_SECRET', default='')
        refresh_token    = config('GOOGLE_REFRESH_TOKEN', default='')

        if not folder:
            self.stderr.write(self.style.ERROR('Falta variável: GOOGLE_DRIVE_FOLDER_ID'))
            return

        if sa_json:
            service = self._drive_service_sa(sa_json)
        elif all([client_id, client_secret, refresh_token]):
            service = self._drive_service(client_id, client_secret, refresh_token)
        else:
            self.stderr.write(self.style.ERROR(
                'Faltam credenciais Google. Configure GOOGLE_SERVICE_ACCOUNT_JSON '
                'ou GOOGLE_CLIENT_ID + GOOGLE_CLIENT_SECRET + GOOGLE_REFRESH_TOKEN'
            ))
            return

        from datetime import datetime, timezone
        ts     = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        erros  = []

        apenas_db    = options['apenas_db']
        apenas_midia = options['apenas_midia']
        apenas_excel = options['apenas_excel']
        manter       = options['manter']

        # ── 1. Banco de dados (Django dumpdata → JSON comprimido) ──
        if not apenas_midia:
            nome_db = f'{PREFIXO_DB}{ts}.json.gz'
            self.stdout.write(f'[1/2] Exportando banco → {nome_db} …')
            try:
                buf = self._dump_db()
                self._upload(service, folder, nome_db, buf, 'application/gzip')
                self.stdout.write(self.style.SUCCESS(f'    ✓ Banco enviado: {nome_db}'))
                self._purgar_antigos(service, folder, PREFIXO_DB, manter)
            except Exception as e:
                msg = f'Falha no backup do banco: {e}'
                self.stderr.write(self.style.ERROR(f'    ✗ {msg}'))
                logger.error(msg)
                erros.append(msg)

        # ── 2. Mídia ──
        if not apenas_db:
            nome_media = f'{PREFIXO_MEDIA}{ts}.tar.gz'
            self.stdout.write(f'[2/2] Compactando mídia → {nome_media} …')
            try:
                buf = self._dump_media()
                self._upload(service, folder, nome_media, buf, 'application/gzip')
                self.stdout.write(self.style.SUCCESS(f'    ✓ Mídia enviada: {nome_media}'))
                self._purgar_antigos(service, folder, PREFIXO_MEDIA, manter)
            except Exception as e:
                msg = f'Falha no backup da mídia: {e}'
                self.stderr.write(self.style.ERROR(f'    ✗ {msg}'))
                logger.error(msg)
                erros.append(msg)

        # ── 3. Excel com todos os dados ──
        if not apenas_db and not apenas_midia:
            nome_excel = f'{PREFIXO_EXCEL}{ts}.xlsx'
            self.stdout.write(f'[3/3] Gerando Excel → {nome_excel} …')
            try:
                buf = self._gerar_excel()
                self._upload(service, folder, nome_excel,
                             buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.stdout.write(self.style.SUCCESS(f'    ✓ Excel enviado: {nome_excel}'))
                self._purgar_antigos(service, folder, PREFIXO_EXCEL, manter)
            except Exception as e:
                msg = f'Falha no Excel: {e}'
                self.stderr.write(self.style.ERROR(f'    ✗ {msg}'))
                logger.error(msg)
                erros.append(msg)

        if erros:
            self.stderr.write(self.style.ERROR(f'Backup concluído COM ERROS: {len(erros)} falha(s).'))
        else:
            self.stdout.write(self.style.SUCCESS('Backup concluído com sucesso!'))

    # ──────────────────────────────────────────────
    def _drive_service_sa(self, sa_json_str):
        """Autenticação via Service Account JSON."""
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        sa_info = json.loads(sa_json_str)
        creds = service_account.Credentials.from_service_account_info(
            sa_info,
            scopes=['https://www.googleapis.com/auth/drive'],
        )
        return build('drive', 'v3', credentials=creds)

    def _drive_service(self, client_id, client_secret, refresh_token):
        """Autenticação via OAuth2 (legado)."""
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        creds = Credentials(
            token=None,
            refresh_token=refresh_token,
            client_id=client_id,
            client_secret=client_secret,
            token_uri='https://oauth2.googleapis.com/token',
        )
        return build('drive', 'v3', credentials=creds)

    def _dump_db(self):
        """Django dumpdata → JSON comprimido em memória."""
        buf_str = io.StringIO()
        call_command(
            'dumpdata',
            '--natural-foreign',
            '--natural-primary',
            '--exclude=contenttypes',
            '--exclude=auth.permission',
            stdout=buf_str,
            verbosity=0,
        )
        buf_gz = io.BytesIO()
        with gzip.GzipFile(fileobj=buf_gz, mode='wb') as gz:
            gz.write(buf_str.getvalue().encode('utf-8'))
        buf_gz.seek(0)
        return buf_gz

    def _dump_media(self):
        """Compacta /app/media em tarball gzip em memória."""
        media_root = getattr(settings, 'MEDIA_ROOT', '/app/media')
        buf = io.BytesIO()
        with tarfile.open(fileobj=buf, mode='w:gz') as tar:
            if os.path.exists(media_root):
                tar.add(media_root, arcname='media')
        buf.seek(0)
        return buf

    def _upload(self, service, folder_id, nome, buf, mime):
        from googleapiclient.http import MediaIoBaseUpload
        meta  = {'name': nome, 'parents': [folder_id]}
        media = MediaIoBaseUpload(buf, mimetype=mime, resumable=True)
        service.files().create(
            body=meta, media_body=media, fields='id',
            supportsAllDrives=True,
        ).execute()

    def _gerar_excel(self):
        """Gera Excel com 4 abas: OS, Boletins, Cadastros, Patrimonial."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from cadastros.models import (
            OrdemServico, BoletimMedicao, Agente, Viatura, Cliente,
            FuncionarioPatrimonial,
        )

        wb = openpyxl.Workbook()
        azul  = PatternFill('solid', fgColor='1F4E79')
        fonte = Font(bold=True, color='FFFFFF', size=10)
        al    = Alignment(horizontal='center', vertical='center')

        def cabecalho(ws, cols):
            ws.append(cols)
            for cell in ws[1]:
                cell.fill = azul
                cell.font = fonte
                cell.alignment = al
            ws.row_dimensions[1].height = 20

        def fmt(v):
            if v is None:
                return ''
            if hasattr(v, 'strftime'):
                return v.strftime('%d/%m/%Y %H:%M') if hasattr(v, 'hour') else v.strftime('%d/%m/%Y')
            return str(v)

        # ── Aba 1: Ordens de Serviço ──────────────────────────
        ws1 = wb.active
        ws1.title = 'Ordens de Serviço'
        cabecalho(ws1, [
            'Nº OS', 'Cliente', 'Solicitante', 'Tipo', 'Status',
            'Previsão Início', 'Origem', 'Destino',
            'Agente 1', 'Agente 2', 'Viatura', 'Placa',
            'Início Viagem', 'Chegada Op.', 'Início Op.', 'Término Op.', 'Término Viagem',
            'KM Início', 'KM Término Op.', 'KM Rodado',
        ])
        for os in OrdemServico.objects.select_related('cliente').prefetch_related('operacional').order_by('-numero'):
            op = None
            try:
                op = os.operacional
            except Exception:
                pass
            ws1.append([
                os.numero,
                os.cliente.razao_social if os.cliente else '',
                os.solicitante,
                os.get_tipo_viagem_display(),
                os.get_status_display(),
                fmt(os.previsao_inicio),
                f'{os.cidade_origem}/{os.uf_origem}' if os.cidade_origem else '',
                f'{os.cidade_destino}/{os.uf_destino}' if os.cidade_destino else '',
                os.snap_agente1_nome,
                os.snap_agente2_nome,
                os.snap_viatura_modelo,
                os.snap_viatura_placa,
                fmt(op.inicio_viagem) if op else '',
                fmt(op.chegada_operacao) if op else '',
                fmt(op.inicio_operacao) if op else '',
                fmt(op.termino_operacao) if op else '',
                fmt(op.termino_viagem) if op else '',
                op.km_inicio_viagem if op else '',
                op.km_termino_operacao if op else '',
                op.km_total if op else '',
            ])

        # ── Aba 2: Boletins de Medição ────────────────────────
        ws2 = wb.create_sheet('Boletins')
        cabecalho(ws2, [
            'Nº OS', 'Cliente', 'Data OS', 'Tabela de Preço', 'Status',
            'Horas Realizadas', 'KM Realizado', 'Horas Excedentes', 'KM Excedente',
            'Valor Escolta', 'Excedente KM', 'Excedente Hora',
            'Pedágio', 'Acréscimo', 'Desconto', 'Valor Total', 'Nº Nota',
        ])
        for b in BoletimMedicao.objects.select_related('os__cliente', 'tabela_preco').order_by('-os__numero'):
            ws2.append([
                b.os.numero,
                b.os.cliente.razao_social if b.os.cliente else '',
                fmt(b.os.previsao_inicio),
                b.tabela_preco.nome if b.tabela_preco else '',
                b.get_status_display() if hasattr(b, 'get_status_display') else b.status,
                str(b.horas_realizadas),
                b.km_realizado,
                str(b.horas_excedentes),
                b.km_excedente,
                float(b.valor_escolta),
                float(b.valor_excedente_km),
                float(b.valor_excedente_hora),
                float(b.valor_pedagio),
                float(b.acrescimo),
                float(b.desconto),
                float(b.valor_total),
                b.numero_nota or '',
            ])

        # ── Aba 3: Cadastros ──────────────────────────────────
        ws3 = wb.create_sheet('Agentes')
        cabecalho(ws3, [
            'Nome', 'CPF', 'RG', 'Telefone', 'Função', 'Status',
            'CNH', 'Val. CNH', 'CNV', 'Val. CNV', 'Endereço',
        ])
        for a in Agente.objects.order_by('nome'):
            ws3.append([
                a.nome, a.cpf, a.rg, a.telefone,
                a.get_funcao_display() if hasattr(a, 'get_funcao_display') else a.funcao,
                a.get_status_display() if hasattr(a, 'get_status_display') else a.status,
                a.cnh, fmt(a.val_cnh) if hasattr(a, 'val_cnh') else '',
                a.cnv, fmt(a.val_cnv) if hasattr(a, 'val_cnv') else '',
                a.endereco or '',
            ])

        ws3b = wb.create_sheet('Viaturas')
        cabecalho(ws3b, ['Placa', 'Modelo', 'Cor', 'Frota', 'MCT ID', 'Status', 'Renavam', 'Chassi'])
        for v in Viatura.objects.order_by('placa'):
            ws3b.append([
                v.placa, v.marca_modelo, v.cor, v.frota, v.mct_id or '',
                v.get_status_display() if hasattr(v, 'get_status_display') else v.status,
                v.renavam or '', v.chassi or '',
            ])

        ws3c = wb.create_sheet('Clientes')
        cabecalho(ws3c, ['Razão Social', 'Nome Fantasia', 'CNPJ', 'Ativo'])
        for c in Cliente.objects.order_by('razao_social'):
            ws3c.append([
                c.razao_social, c.nome_fantasia or '', c.cnpj,
                'Sim' if c.ativo else 'Não',
            ])

        # ── Aba 4: Patrimonial ────────────────────────────────
        ws4 = wb.create_sheet('Patrimonial')
        cabecalho(ws4, [
            'Empresa', 'Nome', 'CPF', 'RG', 'Cargo', 'Status',
            'CNH', 'Val. CNH', 'CNV', 'Val. CNV',
        ])
        for f in FuncionarioPatrimonial.objects.order_by('empresa', 'nome'):
            ws4.append([
                f.get_empresa_display() if hasattr(f, 'get_empresa_display') else f.empresa,
                f.nome, f.cpf, f.rg or '',
                f.cargo or '',
                f.get_status_display() if hasattr(f, 'get_status_display') else getattr(f, 'status', ''),
                f.cnh or '', fmt(f.val_cnh) if getattr(f, 'val_cnh', None) else '',
                f.cnv or '', fmt(f.val_cnv) if getattr(f, 'val_cnv', None) else '',
            ])

        # Ajusta largura das colunas em todas as abas
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _purgar_antigos(self, service, folder_id, prefixo, manter):
        query = (
            f"'{folder_id}' in parents "
            f"and name contains '{prefixo}' "
            f"and trashed = false"
        )
        resp = service.files().list(
            q=query, fields='files(id,name,createdTime)',
            orderBy='createdTime desc',
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        for arq in resp.get('files', [])[manter:]:
            service.files().delete(
                fileId=arq['id'], supportsAllDrives=True
            ).execute()
            self.stdout.write(f'    🗑  Removido: {arq["name"]}')
