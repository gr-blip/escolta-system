# -*- coding: utf-8 -*-
"""
Management command para importar funcionários JRS Facilities a partir de Excel.

Uso:
    python manage.py importar_jrs_facilities
    python manage.py importar_jrs_facilities --arquivo caminho/para/arquivo.xlsx
"""

from datetime import datetime
from django.core.management.base import BaseCommand
from cadastros.models import FuncionarioPatrimonial

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    help = 'Importa funcionários JRS Facilities Ltda a partir do Excel EFETIVO - JRS.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--arquivo',
            type=str,
            default=None,
            help='Caminho do arquivo Excel.',
        )
        parser.add_argument(
            '--json-file',
            type=str,
            default=None,
            help='Caminho para arquivo JSON com os dados (alternativa ao Excel).',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Apenas mostra o que seria importado, sem salvar.',
        )

    def handle(self, *args, **options):
        import json as json_mod

        arquivo = options['arquivo']
        json_file = options.get('json_file')
        dry_run = options['dry_run']

        rows = []

        if json_file:
            self.stdout.write(self.style.NOTICE(f'Lendo JSON: {json_file}'))
            with open(json_file, 'r', encoding='utf-8') as f:
                rows = json_mod.load(f)
        elif arquivo:
            if openpyxl is None:
                self.stderr.write(self.style.ERROR('openpyxl não instalado.'))
                return
            self.stdout.write(self.style.NOTICE(f'Lendo Excel: {arquivo}'))
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
            for row_num in range(6, ws.max_row + 1):
                cpf_raw = ws.cell(row=row_num, column=2).value
                if not cpf_raw:
                    continue
                cpf_limpo = str(cpf_raw).strip().replace('.', '').replace('-', '').replace('/', '')
                if len(cpf_limpo) == 11:
                    cpf_formatado = f'{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}'
                else:
                    cpf_formatado = cpf_limpo
                nome = str(ws.cell(row=row_num, column=3).value or '').strip()
                nasc = ws.cell(row=row_num, column=4).value
                cargo = str(ws.cell(row=row_num, column=5).value or '').strip()
                mae = str(ws.cell(row=row_num, column=6).value or '').strip()
                data_nasc = ''
                if nasc:
                    if hasattr(nasc, 'strftime'):
                        data_nasc = nasc.strftime('%Y-%m-%d')
                    else:
                        data_nasc = str(nasc)
                rows.append({'nome': nome, 'cpf': cpf_formatado, 'cargo': cargo, 'mae': mae, 'nasc': data_nasc})
        else:
            self.stderr.write(self.style.ERROR('Forneça --arquivo ou --json-file'))
            return

        importados = 0
        pulados = 0

        for row in rows:
            nome = row.get('nome', '').strip()
            cpf = row.get('cpf', '').strip()
            cargo = row.get('cargo', '').strip()
            mae = row.get('mae', '').strip()
            nasc = row.get('nasc', '').strip()

            if not nome or not cpf:
                pulados += 1
                continue

            data_nascimento = None
            if nasc:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        data_nascimento = datetime.strptime(nasc, fmt).date()
                        break
                    except ValueError:
                        continue

            if FuncionarioPatrimonial.objects.filter(cpf=cpf).exists():
                self.stdout.write(self.style.WARNING(f'  CPF {cpf} já existe ({nome}), pulando'))
                pulados += 1
                continue

            if dry_run:
                self.stdout.write(f'  [DRY-RUN] {nome} | CPF: {cpf} | Cargo: {cargo}')
                importados += 1
                continue

            FuncionarioPatrimonial.objects.create(
                empresa='jrs_facilities',
                nome=nome,
                cpf=cpf,
                cargo=cargo,
                nome_mae=mae,
                data_nascimento=data_nascimento,
                status='ativo',
            )
            importados += 1

        if dry_run:
            self.stdout.write(self.style.NOTICE(f'\n[DRY-RUN] {importados} seriam importados, {pulados} pulados.'))
        else:
            self.stdout.write(self.style.SUCCESS(f'\nImportação concluída: {importados} importados, {pulados} pulados.'))
