"""
cadastros/boletim_export.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Boletim de Medição — PDF (A3 landscape) e XLSX
Formato conforme modelo aprovado pelo cliente.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import io
from datetime import datetime

# ─── Identidade visual ──────────────────────────────────────────────────────
EMPRESA_NOME = "JR SEGURANÇA E VIGILÂNCIA PATRIMONIAL LTDA – DEPTO. DE ESCOLTA ARMADA"

# Cores (hex sem '#')
_AZUL_HDR   = "1A3A5C"   # cabeçalho título
_AZUL_COL   = "1F4E79"   # header das colunas
_BRANCO     = "FFFFFF"
_CINZA_LIN  = "F2F5F9"   # linhas pares
_ROSA       = "FFB3BA"   # col 9 – Data Hora Encerramento
_ROSA_CLR   = "FFD6DA"   # dados col 9
_AMARELO    = "FFFF00"   # col 10 – Total de Horas
_AMAR_CLR   = "FFFACD"   # dados col 10
_VERMELHO   = "FF0000"   # header col 22 – TOTAL OS
_VERM_CLR   = "FFE0E0"   # dados col 22
_LARANJA    = "FF8C00"   # header cols 25 e 28 – TOTAL km/h
_LARA_CLR   = "FFF0D0"   # dados cols 25 e 28
_VERDE_TOT  = "1A6B35"   # linha VALOR BRUTO
_FUNDO_TOT  = "D6E4F0"   # linha de totais
_CINZA_ROD  = "888888"

# ─── Definição das 33 colunas ───────────────────────────────────────────────
# (título, largura_pdf_rel, largura_xlsx_chars, alinhamento xlsx: L/C/R)
_COLUNAS = [
    # Bloco 1 — Identificação
    ("Ordem\nServiço",                   6,  10, 'C'),
    ("Cliente",                         14,  18, 'L'),
    ("Prestador\nServiço",              10,  14, 'L'),
    ("Placa",                            5,   8, 'C'),
    ("Veículo\nEscoltado",               8,  10, 'C'),
    # Bloco 2 — Datas
    ("Data Hora\nAgendamento VTR",      11,  18, 'C'),
    ("Data Hora\nChegada VTR",          11,  18, 'C'),
    ("Data Hora\nInício",               11,  18, 'C'),
    ("Data Hora\nEncerramento",         11,  18, 'C'),   # col 9 → ROSA
    ("Total de\nHoras",                  7,   9, 'C'),   # col 10 → AMARELO
    # Bloco 3 — Hodômetros
    ("Base\nKm",                         6,   9, 'R'),
    ("Hodômetro\nInicial",               7,  10, 'R'),
    ("Hodômetro\nFinal",                 7,  10, 'R'),
    ("Total\nKm",                        5,   8, 'R'),
    # Bloco 4 — Origem / Destino
    ("Cidade\nOrigem",                  10,  14, 'L'),
    ("UF",                               3,   4, 'C'),
    ("Cidade\nDestino",                 10,  14, 'L'),
    ("UF",                               3,   4, 'C'),
    # Bloco 5 — Franquia
    ("Valor Franquia\nContratada (R$)",  9,  13, 'R'),
    ("Franquia\nContratada de Km",       7,  10, 'R'),
    ("Franquia\nContratada de HORAS",    7,  10, 'C'),
    # Bloco 6 — KM Excedente
    ("TOTAL\nOS",                        6,   9, 'R'),   # col 22 → VERMELHO
    ("KM\nexcedente",                    5,   8, 'R'),
    ("$ POR km\nexcedente",              6,   9, 'R'),
    ("TOTAL",                            6,   9, 'R'),   # col 25 → LARANJA
    # Bloco 7 — Horas Excedentes
    ("Horas\nexcedentes",                6,   9, 'C'),
    ("$ POR\nHoras Excedente",           6,   9, 'R'),
    ("TOTAL",                            6,   9, 'R'),   # col 28 → LARANJA
    # Bloco 8 — Encerramento
    ("Despesas\nExtra Franquia",         7,  11, 'R'),
    ("Fechamento\nCOM / SEM",            8,  12, 'C'),
    ("Data de\nPagamento",               8,  12, 'C'),
    ("NOTA\nFISCAL",                     6,   9, 'C'),
    ("TOTAL",                            8,  11, 'R'),   # col 33 — TOTAL GERAL
]

# Índices 0-based com cor especial
_IDX_ROSA    = 8   # Data Hora Encerramento
_IDX_AMAR    = 9   # Total de Horas
_IDX_VERM    = 21  # TOTAL OS
_IDX_LARA1   = 24  # TOTAL km
_IDX_LARA2   = 27  # TOTAL horas
_IDX_TOTAL   = 32  # TOTAL geral

# Índices de colunas monetárias (R$) — 0-based
_IDX_MOEDA = {18, 21, 23, 24, 26, 27, 28, 32}


# ─── Helpers ────────────────────────────────────────────────────────────────

def _fmt_brl(valor) -> str:
    try:
        v = float(valor)
    except (TypeError, ValueError):
        v = 0.0
    if v == 0:
        return "0,00"
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _missao_to_row(m: dict) -> list:
    """Converte dict da missão para lista de 33 valores (mesma ordem de _COLUNAS)."""
    return [
        m.get('os',          ''),    # 0
        m.get('cliente',     ''),    # 1
        m.get('prestador',   ''),    # 2
        m.get('placa',       ''),    # 3
        m.get('escoltado',   ''),    # 4
        m.get('agendamento', ''),    # 5
        m.get('chegada_vtr', ''),    # 6
        m.get('inicio_op',   ''),    # 7
        m.get('termino_op',  ''),    # 8  ← ROSA
        m.get('total_h',     ''),    # 9  ← AMARELO
        m.get('base_km',     ''),    # 10
        m.get('hod_ini',     ''),    # 11
        m.get('hod_fim',     ''),    # 12
        m.get('total_km',    0),     # 13
        m.get('cidade_ori',  ''),    # 14
        m.get('uf_ori',      ''),    # 15
        m.get('cidade_dst',  ''),    # 16
        m.get('uf_dst',      ''),    # 17
        m.get('valor_franq', 0.0),   # 18
        m.get('franq_km',    0),     # 19
        m.get('franq_horas', ''),    # 20
        m.get('total_os',    0.0),   # 21 ← VERMELHO
        m.get('exced_km',    0),     # 22
        m.get('taxa_km',     0.0),   # 23
        m.get('subtotal_km', 0.0),   # 24 ← LARANJA
        m.get('exced_h',     ''),    # 25
        m.get('taxa_hora',   0.0),   # 26
        m.get('subtotal_h',  0.0),   # 27 ← LARANJA
        m.get('despesas',    0.0),   # 28
        m.get('fechamento',  ''),    # 29
        m.get('dt_pagamento',''),    # 30
        m.get('nota_fiscal', ''),    # 31
        m.get('total',       0.0),   # 32 ← TOTAL
    ]


# ─── PDF (A3 landscape) ─────────────────────────────────────────────────────

def gerar_pdf_bytes(cliente: str, periodo: str, missoes: list, totais: dict) -> bytes:
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    c_azul   = colors.HexColor(f"#{_AZUL_HDR}")
    c_col    = colors.HexColor(f"#{_AZUL_COL}")
    c_cinza  = colors.HexColor(f"#{_CINZA_LIN}")
    c_rosa_h = colors.HexColor(f"#{_ROSA}")
    c_rosa_d = colors.HexColor(f"#{_ROSA_CLR}")
    c_amar_h = colors.HexColor(f"#{_AMARELO}")
    c_amar_d = colors.HexColor(f"#{_AMAR_CLR}")
    c_verm_h = colors.HexColor(f"#{_VERMELHO}")
    c_verm_d = colors.HexColor(f"#{_VERM_CLR}")
    c_lara_h = colors.HexColor(f"#{_LARANJA}")
    c_lara_d = colors.HexColor(f"#{_LARA_CLR}")
    c_vtot   = colors.HexColor(f"#{_VERDE_TOT}")
    c_ftot   = colors.HexColor(f"#{_FUNDO_TOT}")
    c_bord   = colors.HexColor('#B0BBC8')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                            leftMargin=8*mm, rightMargin=8*mm,
                            topMargin=8*mm, bottomMargin=8*mm)
    W = landscape(A3)[0] - 16*mm

    def _st(name, size=6, bold=False, align=TA_LEFT, color=colors.black, leading=7):
        return ParagraphStyle(name, fontSize=size,
                              fontName='Helvetica-Bold' if bold else 'Helvetica',
                              alignment=align, textColor=color, leading=leading)

    st_titulo = _st('tt', 13, True, TA_CENTER, c_azul, 16)
    st_emp    = _st('te',  7, True, TA_RIGHT,  c_azul,  9)
    st_cli    = _st('tc',  8, True, TA_LEFT,   c_azul)
    st_hdr    = _st('th',  5, True, TA_CENTER, colors.white, 6)
    st_L      = _st('tL',  5, False, TA_LEFT,  colors.black, 6)
    st_R      = _st('tR',  5, False, TA_RIGHT, colors.black, 6)
    st_C      = _st('tC',  5, False, TA_CENTER,colors.black, 6)
    st_bR     = _st('bR',  5, True,  TA_RIGHT, c_vtot, 6)
    st_bL     = _st('bL',  5, True,  TA_LEFT,  c_vtot, 6)
    st_rod    = _st('rd',  6, False, TA_CENTER, colors.HexColor(f"#{_CINZA_ROD}"))

    story = []
    data_ger = datetime.now().strftime("%d/%m/%Y %H:%M")
    n_regs   = totais.get('missoes', len(missoes))

    # Cabeçalho
    hdr = [[
        Paragraph("<b>BOLETIM DE MEDIÇÃO</b>", st_titulo),
        Paragraph(f"<b>{EMPRESA_NOME}</b><br/>Período: {periodo} &nbsp;|&nbsp; {n_regs:03d} registros", st_emp),
    ]]
    t_hdr = Table(hdr, colWidths=[W * 0.5, W * 0.5])
    t_hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#EBF2FA')),
        ('LINEBELOW', (0,0), (-1,-1), 1, c_azul),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(f"<b>Cliente: {cliente}</b>", st_cli))
    story.append(Spacer(1, 2*mm))

    # Larguras proporcionais
    raw_w  = [c[1] for c in _COLUNAS]
    col_w  = [r * (W / sum(raw_w)) for r in raw_w]

    # Cabeçalho das colunas
    def _hdr_cell(i, col):
        title  = col[0]
        # cor do header baseada no índice
        if i == _IDX_ROSA:
            bg = c_rosa_h; fc = colors.black
        elif i == _IDX_AMAR:
            bg = c_amar_h; fc = colors.black
        elif i == _IDX_VERM:
            bg = c_verm_h; fc = colors.white
        elif i in (_IDX_LARA1, _IDX_LARA2):
            bg = c_lara_h; fc = colors.white
        else:
            bg = c_col; fc = colors.white
        return Paragraph(f"<b>{title}</b>",
                         _st(f'h{i}', 4.5, True, TA_CENTER, fc, 5.5))

    header_row = [_hdr_cell(i, c) for i, c in enumerate(_COLUNAS)]
    rows = [header_row]

    # Linhas de dados
    _aligns = [c[3] for c in _COLUNAS]
    for ri, m in enumerate(missoes):
        vals = _missao_to_row(m)
        row  = []
        for i, v in enumerate(vals):
            al = _aligns[i]
            st = st_R if al == 'R' else (st_C if al == 'C' else st_L)
            if i in _IDX_MOEDA:
                cell = Paragraph(_fmt_brl(v), st_R)
            elif i == _IDX_TOTAL:
                cell = Paragraph(f"<b>{_fmt_brl(v)}</b>",
                                 _st(f'd{i}', 5, True, TA_RIGHT, c_vtot, 6))
            else:
                cell = Paragraph(str(v) if v != '' else '—', st)
            row.append(cell)
        rows.append(row)

    # Linha de totais gerais
    def _tv(k, moeda=True):
        v = totais.get(k, 0)
        return Paragraph(f"<b>{_fmt_brl(v) if moeda else v}</b>", st_bR)

    tot_vals = [''] * len(_COLUNAS)
    tot_vals[0]  = Paragraph(f"<b>TOTAL — {n_regs:03d} OS</b>", st_bL)
    tot_vals[9]  = Paragraph(f"<b>{totais.get('total_h','')}</b>",   st_bR)
    tot_vals[13] = Paragraph(f"<b>{totais.get('total_km','')}</b>",  st_bR)
    tot_vals[21] = _tv('subtotal_km')
    tot_vals[22] = Paragraph(f"<b>{totais.get('exced_km','')}</b>",  st_bR)
    tot_vals[24] = _tv('subtotal_km')
    tot_vals[25] = Paragraph(f"<b>{totais.get('exced_h','')}</b>",   st_bR)
    tot_vals[27] = _tv('subtotal_h')
    tot_vals[28] = _tv('despesas')
    tot_vals[32] = _tv('total')
    rows.append(tot_vals)

    # VALOR BRUTO
    vb_row = [''] * len(_COLUNAS)
    vb_row[31] = Paragraph("<b>VALOR BRUTO</b>",
                           _st('vbl', 6, True, TA_RIGHT, c_vtot, 7))
    vb_row[32] = Paragraph(f"<b>R$ {_fmt_brl(totais.get('total', 0))}</b>",
                           _st('vbv', 6, True, TA_RIGHT, c_vtot, 7))
    rows.append(vb_row)

    tabela = Table(rows, colWidths=col_w, repeatRows=1)
    n_data = len(rows) - 3  # rows sem header, totais e valor bruto

    cmds = [
        ('GRID',         (0,0),  (-1,-2), 0.3, c_bord),
        ('VALIGN',       (0,0),  (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0),  (-1,-1), 1),
        ('BOTTOMPADDING',(0,0),  (-1,-1), 1),
        ('LEFTPADDING',  (0,0),  (-1,-1), 2),
        ('RIGHTPADDING', (0,0),  (-1,-1), 2),
        # Linha de totais
        ('BACKGROUND',   (0,-2), (-1,-2), c_ftot),
        ('LINEABOVE',    (0,-2), (-1,-2), 1.2, c_azul),
        # Linha VALOR BRUTO
        ('BACKGROUND',   (0,-1), (-1,-1), colors.white),
        ('LINEABOVE',    (0,-1), (-1,-1), 0.5, c_bord),
        # Cores especiais no header (row 0)
        ('BACKGROUND',   (_IDX_ROSA,  0), (_IDX_ROSA,  0), c_rosa_h),
        ('BACKGROUND',   (_IDX_AMAR,  0), (_IDX_AMAR,  0), c_amar_h),
        ('BACKGROUND',   (_IDX_VERM,  0), (_IDX_VERM,  0), c_verm_h),
        ('BACKGROUND',   (_IDX_LARA1, 0), (_IDX_LARA1, 0), c_lara_h),
        ('BACKGROUND',   (_IDX_LARA2, 0), (_IDX_LARA2, 0), c_lara_h),
        # Cor padrão das outras colunas no header
        ('BACKGROUND',   (0, 0), (-1, 0), c_col),
        # Reaplicar cores especiais por cima (ORDER matters em ReportLab)
        ('BACKGROUND',   (_IDX_ROSA,  0), (_IDX_ROSA,  0), c_rosa_h),
        ('BACKGROUND',   (_IDX_AMAR,  0), (_IDX_AMAR,  0), c_amar_h),
        ('BACKGROUND',   (_IDX_VERM,  0), (_IDX_VERM,  0), c_verm_h),
        ('BACKGROUND',   (_IDX_LARA1, 0), (_IDX_LARA1, 0), c_lara_h),
        ('BACKGROUND',   (_IDX_LARA2, 0), (_IDX_LARA2, 0), c_lara_h),
    ]

    # Cores especiais nas linhas de dados
    for ri in range(1, n_data + 1):
        base_bg = c_cinza if ri % 2 == 0 else colors.white
        cmds.append(('BACKGROUND', (0, ri), (-1, ri), base_bg))
        cmds.append(('BACKGROUND', (_IDX_ROSA,  ri), (_IDX_ROSA,  ri), c_rosa_d))
        cmds.append(('BACKGROUND', (_IDX_AMAR,  ri), (_IDX_AMAR,  ri), c_amar_d))
        cmds.append(('BACKGROUND', (_IDX_VERM,  ri), (_IDX_VERM,  ri), c_verm_d))
        cmds.append(('BACKGROUND', (_IDX_LARA1, ri), (_IDX_LARA1, ri), c_lara_d))
        cmds.append(('BACKGROUND', (_IDX_LARA2, ri), (_IDX_LARA2, ri), c_lara_d))

    tabela.setStyle(TableStyle(cmds))
    story.append(tabela)
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Gerado em: {data_ger} | {EMPRESA_NOME}", st_rod))

    doc.build(story)
    return buf.getvalue()


# ─── XLSX ────────────────────────────────────────────────────────────────────

def gerar_xlsx_bytes(cliente: str, periodo: str, missoes: list, totais: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Boletim de Medição"

    N  = len(_COLUNAS)
    LC = get_column_letter(N)

    def _fill(hex6):
        return PatternFill('solid', fgColor=hex6)

    def _borda(tipo='thin', cor='B0BBC8'):
        s = Side(style=tipo, color=cor)
        return Border(left=s, right=s, top=s, bottom=s)

    def _font(bold=False, size=8, color='000000', italic=False):
        return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)

    def _align(h='left', wrap=False):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

    data_ger = datetime.now().strftime("%d/%m/%Y %H:%M")
    n_regs   = totais.get('missoes', len(missoes))

    # ── Linha 1: Título ────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{LC}1')
    ws['A1'] = "BOLETIM DE MEDIÇÃO"
    ws['A1'].font      = _font(bold=True, size=14, color=_AZUL_HDR)
    ws['A1'].alignment = _align('center')
    ws['A1'].fill      = _fill('EBF2FA')
    ws.row_dimensions[1].height = 26

    # ── Linha 2: Empresa + Período ─────────────────────────────────────────
    mid = N // 2
    ws.merge_cells(f'A2:{get_column_letter(mid)}2')
    ws['A2'] = EMPRESA_NOME
    ws['A2'].font      = _font(bold=True, size=8, color=_AZUL_HDR)
    ws['A2'].alignment = _align('left')
    ws['A2'].fill      = _fill('EBF2FA')

    c2 = get_column_letter(mid + 1)
    ws.merge_cells(f'{c2}2:{LC}2')
    ws[f'{c2}2'] = f"Período: {periodo}  |  {n_regs:03d} registros"
    ws[f'{c2}2'].font      = _font(size=8, color=_AZUL_HDR)
    ws[f'{c2}2'].alignment = _align('right')
    ws[f'{c2}2'].fill      = _fill('EBF2FA')
    ws.row_dimensions[2].height = 18

    # ── Linha 3: Cliente ───────────────────────────────────────────────────
    ws.merge_cells(f'A3:{LC}3')
    ws['A3'] = f"Cliente: {cliente}"
    ws['A3'].font      = _font(bold=True, size=9, color=_AZUL_HDR)
    ws['A3'].alignment = _align('left')
    ws['A3'].fill      = _fill('EBF2FA')
    ws.row_dimensions[3].height = 18

    # ── Linha 4: Cabeçalho das colunas ────────────────────────────────────
    HDR_ROW = 4
    # Mapa de cores dos headers especiais (1-based)
    _hdr_bg = {
        _IDX_ROSA  + 1: (_ROSA,     '000000'),
        _IDX_AMAR  + 1: (_AMARELO,  '000000'),
        _IDX_VERM  + 1: (_VERMELHO, _BRANCO),
        _IDX_LARA1 + 1: (_LARANJA,  _BRANCO),
        _IDX_LARA2 + 1: (_LARANJA,  _BRANCO),
    }

    for ci, (nome, _, col_w, _al) in enumerate(_COLUNAS, start=1):
        cel = ws.cell(row=HDR_ROW, column=ci, value=nome)
        bg, fc = _hdr_bg.get(ci, (_AZUL_COL, _BRANCO))
        cel.font      = _font(bold=True, size=8, color=fc)
        cel.fill      = _fill(bg)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border    = _borda()
        ws.column_dimensions[get_column_letter(ci)].width = col_w
    ws.row_dimensions[HDR_ROW].height = 34

    # ── Linhas de dados ────────────────────────────────────────────────────
    # Mapa de cores das células de dados especiais (1-based col)
    _dat_bg = {
        _IDX_ROSA  + 1: _ROSA_CLR,
        _IDX_AMAR  + 1: _AMAR_CLR,
        _IDX_VERM  + 1: _VERM_CLR,
        _IDX_LARA1 + 1: _LARA_CLR,
        _IDX_LARA2 + 1: _LARA_CLR,
    }
    _aligns = [c[3] for c in _COLUNAS]
    _moeda1 = {i + 1 for i in _IDX_MOEDA}   # 1-based

    for ri, m in enumerate(missoes, start=HDR_ROW + 1):
        vals = _missao_to_row(m)
        base_bg = _CINZA_LIN if ri % 2 == 0 else _BRANCO
        for ci, val in enumerate(vals, start=1):
            is_total_col = (ci == _IDX_TOTAL + 1)
            # Horas: manter como texto
            is_hora = (ci in {_IDX_AMAR + 1, 26})  # total_h e exced_h
            if is_hora and val == '':
                val = '00:00'
            bg  = _dat_bg.get(ci, base_bg)
            al  = _aligns[ci - 1]
            cel = ws.cell(row=ri, column=ci, value=val)
            if is_hora:
                cel.number_format = '@'
                cel.data_type     = 's'
            cel.font      = _font(bold=is_total_col, size=7,
                                  color=(_VERDE_TOT if is_total_col else '000000'))
            cel.fill      = _fill(bg)
            cel.border    = _borda()
            cel.alignment = _align('right' if al == 'R' else ('center' if al == 'C' else 'left'),
                                   wrap=(ci <= 5))
            if ci in _moeda1 and not is_hora:
                cel.number_format = '#,##0.00'
        ws.row_dimensions[ri].height = 20

    # ── Linha de Totais ────────────────────────────────────────────────────
    TOT_ROW = HDR_ROW + 1 + len(missoes)
    tot_map = {
        1:  f"TOTAL — {n_regs:03d} OS",
        10: totais.get('total_h', ''),
        14: totais.get('total_km', 0),
        22: totais.get('exced_km', 0),
        25: totais.get('subtotal_km', 0),
        23: totais.get('exced_km', 0),
        26: totais.get('exced_h', ''),
        28: totais.get('subtotal_h', 0),
        29: totais.get('despesas', 0),
        33: totais.get('total', 0),
    }
    for ci in range(1, N + 1):
        val = tot_map.get(ci, '')
        is_h = ci in {10, 26}
        cel = ws.cell(row=TOT_ROW, column=ci, value=val)
        if is_h:
            cel.number_format = '@'
            cel.data_type = 's'
        cel.font   = _font(bold=True, size=8,
                           color=(_VERDE_TOT if ci in _moeda1 else _AZUL_HDR))
        cel.fill   = _fill(_FUNDO_TOT)
        cel.border = _borda('medium', '1A3A5C')
        cel.alignment = _align('right' if ci >= 10 else 'left')
        if ci in _moeda1 and not is_h:
            cel.number_format = '#,##0.00'
    ws.row_dimensions[TOT_ROW].height = 22

    # ── VALOR BRUTO ────────────────────────────────────────────────────────
    VB_ROW = TOT_ROW + 1
    ws.merge_cells(f'A{VB_ROW}:{get_column_letter(N-1)}{VB_ROW}')
    ws[f'A{VB_ROW}'] = "VALOR BRUTO"
    ws[f'A{VB_ROW}'].font      = _font(bold=True, size=10, color=_BRANCO)
    ws[f'A{VB_ROW}'].fill      = _fill(_VERDE_TOT)
    ws[f'A{VB_ROW}'].alignment = _align('right')

    vb_cel = ws.cell(row=VB_ROW, column=N, value=totais.get('total', 0))
    vb_cel.font          = _font(bold=True, size=10, color=_BRANCO)
    vb_cel.fill          = _fill(_VERDE_TOT)
    vb_cel.alignment     = _align('right')
    vb_cel.number_format = 'R$ #,##0.00'
    ws.row_dimensions[VB_ROW].height = 24

    # ── Rodapé ────────────────────────────────────────────────────────────
    ROD = VB_ROW + 1
    ws.merge_cells(f'A{ROD}:{LC}{ROD}')
    ws[f'A{ROD}'] = f"Gerado em: {data_ger}  |  {EMPRESA_NOME}"
    ws[f'A{ROD}'].font      = _font(size=7, italic=True, color=_CINZA_ROD)
    ws[f'A{ROD}'].alignment = _align('center')

    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
